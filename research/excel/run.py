import openpyxl
import re
import calendar

_values_to_eliminate =['%Change','Thousand Metric','Month','Date']
valid_sheets = ["T1","T2","T3","T4","T5","T6","T7","T8"]
loader = {}
collections_mapper =[]


'''
 S1 : Check for headers to be present in various rows in the file, if yes, split by the headers
 S2 : Check for Sub-Tables inside the main table eg : Sheet-T8 , oil1114.xlsx
'''

def read_excel(excelfull_path):


	sheet = openpyxl.load_workbook(excelfull_path)

	for sheets_by_name in valid_sheets:

		ws = sheet.get_sheet_by_name(sheets_by_name)	
		
		'''
			S0: Load all the Sheet Values into a Container for Processing
		'''
		get_rows_val = (load_values_container(container = [] , iter_rows =ws.iter_rows()))

		'''
			S1: Fetch the Header information from the Tables
		'''

		header_value , mx_mn ,len_max , len_min, previous_row,next_row =  determine_headers(get_rows_val)
		original_header_value = str(''.join(header_value))

		'''
			S2: Modify the Header to accept missing cells such as one below the header
			This function will get executed only if there are any such identifiers below the headers that
			are neither of len_max or close to it..

			For eg ::

				2012	2013	1Q2013	2Q2013	3Q2013	4Q2013	Jan2014	Current	Year to -> Header
																		Month1	Date2 -> Cells belonging to Header
			Function will add those cells to the Header														
		'''
		
		modified_header_value = modify_header(header_value , mx_mn ,len_max , len_min, previous_row,next_row)

		
		'''
			S2: Iterate over the entire sheet and modify all headers as modified_header_value
		'''

		modified_rows = modify_table_rows(get_rows_val,original_header_value,modified_header_value)

		modified_rows = eliminate_values_container(modified_rows,modified_header_value)
		

		'''
			S1 : Headers are present in more than one location  - Refer Sheet T3 , File -> oil1114.xlsx

			OECD Americas: Imports of Crude Oil, NGL and Refinery Feedstocks 1							
								Thousand metric tons
			Imports from:	2012	2013	1Q2013	2Q2013	3Q2013	4Q2013	Jan2014 --> Header
			Australia	274	143	2	65	76	-	-
			Belgium	2017	2362	501	514	770	577	164
			Canada	127017	135706	34998	31743	33658	35307	12955		
			Denmark	7	62	56	6	-	-	-

			OECD Asia Oceania: Imports of Crude Oil, NGL and Refinery Feedstocks 1							
								Thousand metric tons
			Imports from:	2012	2013	1Q2013	2Q2013	3Q2013	4Q2013	Jan2014
			Australia	3097	3416	769	985	708	954	115
			Belgium	-	-	-	-	-	-	-
			Canada	-	-	-	-	-	-	-
			Denmark	-	-	-	-	-	-	-
			France	-	-	-	-	-	-	-

		'''
		
		get_list_headers_location = check_headers_in_file(modified_header_value,modified_rows)



		'''
			S2 : Check for Sub-Tables - Refer Sheet T8 , File -> oil1114.xlsx

			The Sheet has only one header but the rows are split by keys representing values beneath
			and all the keys belong to the same table and header

												%Change	%Change
				2012	2013	1Q2013	2Q2013	3Q2013	4Q2013	Jan2014	Current	Year to **** --> Header **** 
										Month1	Date2
			OECD Americas **** --> Key **** 
			
			Liquefied Petroleum Gases	20978	20805	4185	7130	6587	2903	1319	16.5	16.5
			Naphtha	15530	16839	4026	4217	4395	4201	1466	1.2	1.2
			Total Gasoline3	431393	442597	105054	110542	112624	114377	37072	3.7	3.7
			
			
			OECD Asia Oceania **** --> Key **** 
			Liquefied Petroleum Gases	6967	7527	1817	1929	2048	1733	662	7.6	7.6
			Naphtha	38992	40102	10535	9174	10334	10059	3377	-7.5	-7.5
			Total Gasoline3	70715	70529	17432	16850	17951	18296	6056	-0.1	-0.1

		'''

		#presence_of_sub_tables = check_presence_of_sub_tables(get_rows_val,len_max , len_min,header_value)

		'''
			Prepare data for the final table
		'''

		mapping = prepare_data_final_table(get_list_headers_location,modified_rows,modified_header_value)



		'''
			Check if the tables have divisions - A table under same header has divisions for each region

			Eg ::

													%Change	%Change
				2012	2013	1Q2013	2Q2013	3Q2013	4Q2013	Jan2014	Current	Year to ---> Header
											Month1	Date2
				Switzerland	 --> Division 1
				Liquefied Petroleum Gases	172	175	52	43	43	37	15	-16.7	-16.7
				Naphtha	13	5	1	1	1	2	-	-100.0	-100.0
				Total Gasoline3	2930	2804	652	718	733	701	212	-1.9	-1.9

				Turkey --> Division 2
				Liquefied Petroleum Gases	3622	3734	805	931	1039	959	244	-9.0	-9.0
				Naphtha	1675	756	151	137	188	280	76	43.4	43.4
				Total Gasoline3	1038	1287	333	330	362	262	52	-61.5	-61.5


		'''

		are_divisions_exist = bool(check_for_divisions_table(mapping))

		# IF the Tables don't have divisions, call the final table loader directly.

		if not bool(are_divisions_exist):

				# For the Tables in the Container
				for table_name, table_content in mapping:

					# For the final table
					filtered_table_content = [items for items in table_content if len(items) == len_max]
					design_table(table_name = table_name,table_headers = table_content[0],table_content=filtered_table_content)
	


def modify_table_rows(get_rows_val,header_value,modified_header_value):

	'''
		Modify the Headers at all location across the sheet
	'''

	for pos,items in enumerate(get_rows_val):

		if header_value in ''.join(items): 

			get_rows_val[pos] = modified_header_value	

	return get_rows_val

def modify_header(header_value , mx_mn ,len_max , len_min, previous_row,next_row):

	'''
		Modify the Table header to include any missing Cells
	'''

	# Get length on the next_row, accordingly map to the headers in a reverse order, e.g. .. if length of next_row is 3, last 3 items of headers will be modified
	get_len_next_row = len(next_row)
	get_len_header = len(header_value)
	next_row = [next_row[-i] for i in range(1,len(next_row)+1)]

	if get_len_next_row > 0 and not get_len_next_row == len_max and not get_len_next_row == len_max - 1:

		for i in range(1,len(next_row)+1):
			try:
				modified_header_value = header_value.__getitem__(-(i))+'_'+next_row.__getitem__(i-1)				
				header_value[get_len_header-i] = modified_header_value
			except Exception :
				continue

	return header_value


def check_for_divisions_table(mapping):


	for table_name,table_content in mapping:

		get_len_items = [len(items) for items in table_content]

		
		'''
			Get records with max and min length
		'''
		

		'''
			Find the indexes of records with min length, consider only records that 
			occur between max values.. For eg : max min max .. say ... 10 1 10 1 10
		'''

		get_index_all_min = [positions for positions,items in enumerate(table_content) if len(items) == min(get_len_items)]
		get_index_all_max = [positions for positions,items in enumerate(table_content) if len(items) == max(get_len_items)]

		
		'''
			Filter the indexes of min positions such that they are not beyond the max boundaries

			For eg :: Values such as these @ tail end of the table shall be eliminated.

			1.  Percentage change over corresponding month of previous year				
			2.  Percentage change over corresponding period (beginning of year to current month) of previous year.				
			3.  Total Gasoline includes Motor Gasoline, Jet Gasoline and Aviation Gasoline 				
			4.  Total Kerosene includes Jet Kerosene and Other Kerosene				
			For country specific notes on data, please see notes in the appendix section				

			Pick values that fall in a valid range ... For eg : max min max .. say ... 10 1 10 1 10

		'''
		get_final_minitems_filtered = {}
		get_final_minitems = [get_final_minitems_filtered.__setitem__(min_items,min_items) for min_items in get_index_all_min for max_items in get_index_all_max if min_items < max_items]


		get_positions_min = sorted(get_final_minitems_filtered.values())				


		for values in fetch_table_based_on_posts(get_positions_min,table_content,mapper = [],apply_re=False):		
			

			# Eliminate records that are != maxlen	

			filtered_table_content = [items for items in values if len(items) == max(get_len_items)]	

			#Form the Table name
			inner_table_name = str(table_name+'_'+''.join(values[0])).strip()
			inner_table_name = inner_table_name.replace(" ","_")

			# Design and output the table
			design_table(inner_table_name,table_content[0],filtered_table_content)



def clear_table_headers(table_header,*args):

	for items in table_header:
		try:
			for values in args:
				if str(values).lower() in str(items).lower():
					table_header.remove(items)
		except Exception as E:
			continue

	return table_header


def design_table(table_name,table_headers,table_content):


	'''
		params : table_name -> Name of the Table
		params : table_headers -> Headers tp map to the Table
		params : table_content -> Entire Table Content
	'''

	# Take the first item from the table_content container and mark it as Product for each item in the Container

	values = [items[1:] for items in table_content]
	products = [items[0] for items in table_content]

	# Remove unwanted charecters from the header
	table_headers = [items.replace(".","").replace("00:","").replace(":00","").replace("00","") for items in table_headers]


	clear_table_headers(table_headers,"Exports","Imports")
	

	#table_headers = [calendar.month_name[int(items.split("-")[-1].replace("0",""))]+"_"+items.split("-")[0] for items in table_headers  if "-" in items]

	prd_id = -1
	for each_prd in products:
		header_counter = 0
		prd_id += 1
		for each_header in table_headers:			
			print table_name,each_prd,each_header,values.__getitem__(prd_id).__getitem__(header_counter),'\n\n'
			loader.__setitem__("Table_name",str(table_name).strip())
			header_counter += 1




def fetch_table_based_on_posts(get_rows_val_split_headers_indexes,get_rows_val,mapper=[],apply_re=True):

			if not mapper == [] : mapper = []	

			for indexid,indexes in enumerate(get_rows_val_split_headers_indexes):

				try:

					'''
						Fetch the Tables from the Excel Sheet - Based on the presence of Headers
					'''
					
					table_content = get_rows_val[indexes:get_rows_val_split_headers_indexes.__getitem__(indexid+1)]					

					if (apply_re): 
						mapper.append((re.sub('[^A-Za-z0-9\.]+', '_', ''.join(get_rows_val[indexes-1])).strip(),table_content))
					else:
						mapper.append(table_content)

				except IndexError:				

					'''
						The last index position of the occurrence of the Headers in the file

						At this point, we had got all the tables from the sheet.
					'''

					table_content = get_rows_val[get_rows_val_split_headers_indexes[-1]:]

					if (apply_re): 
						mapper.append((re.sub('[^A-Za-z0-9\.]+', '_', ''.join(get_rows_val[indexes-1])).strip(),table_content)) 
					else:
						mapper.append(table_content)

			return mapper

def prepare_data_final_table(get_list_headers_location,get_rows_val,header_value,mapper=[]):

	'''
		Form the Final Table
	'''

	if not mapper == [] : mapper = []

	if len(get_list_headers_location) > 0:

		'''
			Step 1: Get the indexes of the headers in the table
		'''

		get_rows_val_split_headers_indexes = [i for i,x in enumerate(get_rows_val) if x == header_value]


		'''
			Step 2: Split the table into multiple based on the indexes
		'''

		return fetch_table_based_on_posts(get_rows_val_split_headers_indexes,get_rows_val,mapper)


def check_presence_of_sub_tables(get_rows_val,len_max , len_min,header_value):
	'''
		Iterate through the table rows and check for sub-tables as described above
	'''
	
	# Check all indexes whose length is 1
	
	get_min_position = [position for position,value in enumerate(get_rows_val) if len(value) == len_min]

	
	# Filter the index whose next index has a length on max ==> len_max

	if get_min_position:

		'''
			Return the indexes of Sub-tables headers, the rows that have only 1 cell and next row contains max entries as
			figured out by len_max
		'''

		return [pos_index for pos_index in get_min_position if not pos_index+1 == len(get_rows_val) and len(get_rows_val.__getitem__(pos_index+1)) == len_max \
			and not get_rows_val.__getitem__(pos_index+1) == header_value]
	
		


def check_headers_in_file(headerinfo,get_rows_val,row_id_info = []):

	if not row_id_info == None: row_id_info = []
	
	for row_id,row_data in enumerate(get_rows_val):
		if headerinfo == row_data:			
			row_id_info.append(row_id)	
	return row_id_info


def load_values_container(container,iter_rows):


	'''
		Get all sheet values into a Container
	'''
	if container is not None:
		container = []

	for row in iter_rows:	

		row_data = [cell.value for cell in row if cell.value is not None]

		if row_data  != []: 	

			new_cell_val = ''
			for each_cell in row_data:				
				try:
					new_cell_val += str(each_cell).strip()+","
					
				except UnicodeEncodeError:
					new_cell_val += re.sub('[^A-Za-z0-9\.]+', '-', each_cell).strip()+','								
			
			container.append(new_cell_val.split(",")[0:-1])						

	return container

def eliminate_values_container(get_rows_val,modified_header_value,elims = {}):

	if not elims is None: elims = {}

	for pos_id,data in enumerate(get_rows_val):
		for values in _values_to_eliminate:
			if not ''.join(modified_header_value).lower() == ''.join(data).lower():
				if values.lower() in ''.join(data).lower():				
					elims[pos_id] = data
		
	for key,values in elims.items():		
		get_rows_val.remove(values)

	return get_rows_val


def determine_headers(iter_rows_data):

	
	lenght_list = [len(items) for items in iter_rows_data]
	len_max , len_min = max(lenght_list),min(lenght_list)

	for pos,items in enumerate(iter_rows_data):		

		if len(items) == len_max or len(items) == len_max - 1:					

			return items,len(items) == len_max,len_max , len_min,iter_rows_data.__getitem__(pos-1),iter_rows_data.__getitem__(pos+1)

def uplod_period(excelfile):

	'''
		params : excelfile - The full path of the excel file

		Identify the Upload period from the file name
	'''


read_excel("OIL1114.xlsx")