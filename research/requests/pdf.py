from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import HTMLConverter
from pdfminer.converter import TextConverter
from pdfminer.pdfparser import PDFParser
from pdfminer.pdfdocument import PDFDocument
import pdfminer.layout
from pdfminer.layout import LAParams,LTTextBox,LTTextLine,LTFigure,LTTextLineHorizontal,LTTextBoxHorizontal
from pdfminer.pdfpage import PDFPage
from pdfminer.converter import PDFPageAggregator
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
import re
import csv
import PyPDF2

from cStringIO import StringIO


import copy
import math
from PyPDF2 import PdfFileReader, PdfFileWriter
import logging
import openpyxl
import requests

logging.basicConfig(level=logging.INFO)

core = logging.getLogger("Caesar")

#searchTerm = "World oil demand in 2017*, mb/d"
searchTerm = "World oil demand in 2017*"
searchTerm2 = "World oil demand in 2016*"
searchTerm3 = "US oil demand, tb/d"

filterterms = ["Note:","Source:","OPEC Secretariat"]
api_key = "iqepqp1fl2t5"

baseurl = 'https://pdftables.com/api?key='+api_key+'&format=xlsx-multiple'
headers = {'content-type': 'multipart/form-data'}


current_search_term = searchTerm

def convert_pdf_to_html(path,htmlfilename):

    rsrcmgr = PDFResourceManager()
    retstr = StringIO()
    codec = 'utf-8'
    laparams = LAParams()
    device = HTMLConverter(rsrcmgr, retstr, codec=codec, laparams=laparams)
    fp = file(path, 'rb')
    print fp, file
    interpreter = PDFPageInterpreter(rsrcmgr, device)
    password = ""
    maxpages = 0 #is for all
    caching = True
    pagenos=set()
    for page in PDFPage.get_pages(fp, pagenos, maxpages=maxpages, password=password,caching=caching, check_extractable=True):
        interpreter.process_page(page)
    fp.close()
    device.close()
    str = retstr.getvalue()
    retstr.close()
    with open(htmlfilename,'wb') as writer:
        writer.write(str)
    return




def _extract_data_from_pdf(pdffile,searchTerm):


    pdffileobj = open(pdffile,'rb')
    
    pdfreader = PyPDF2.PdfFileReader(pdffileobj)
    pdfWriter = PyPDF2.PdfFileWriter()


    for pages in xrange(pdfreader.numPages):

        pageobj = pdfreader.getPage(pages)

        
        xdata = pageobj.extractText().encode('utf-8')
        #xdatan = ''.join(xdata).encode('utf-8')
        tabledata = xdata.split('\t')
        searchdata = tabledata[0].replace('\n','').replace('\r','')

        if searchTerm.lower() in searchdata.lower():
            pdfWriter.addPage(pageobj)           
      
            pdfOutputFile = open(pdffile.split('.')[0]+"_"+str(pages)+'.pdf', 'wb')
            pdfWriter.write(pdfOutputFile)
        
    return pdfOutputFile.name


def _read_pdf(pdffile):

    pdffileobj = open(pdffile,'rb')
    
    pdfreader = PyPDF2.PdfFileReader(pdffileobj)
    pdfWriter = PyPDF2.PdfFileWriter()


    for pages in xrange(pdfreader.numPages):

        pageobj = pdfreader.getPage(pages)        
        xdata = pageobj.extractText().encode('utf-8')
        yield xdata



def pdf_text(fpath,txtfilename):

    with open(fpath,'rb') as fp:
            doc=PDFDocument(PDFParser(fp))
            rsrcmgr=PDFResourceManager()
            retstr=StringIO()
            laparams=LAParams()
            codec='utf-8'
            device=TextConverter(rsrcmgr,retstr,codec=codec,laparams=laparams)
            interpreter=PDFPageInterpreter(rsrcmgr,device)
            lines=""
            for page in PDFPage.create_pages(doc):
                interpreter.process_page(page)
                rstr=retstr.getvalue()
                
                if len(rstr.strip()) >0:
                    lines+="".join(rstr)
            
            with open(txtfilename,'wb') as textWriter :

               textWriter.write(lines)

            return txtfilename
    


def _get_specific_page(pdffile,name,pagID):


    pdffileobj = open(pdffile,'rb')
    
    pdfreader = PyPDF2.PdfFileReader(pdffileobj)
    pdfWriter = PyPDF2.PdfFileWriter()


    #for pages in xrange(pdfreader.numPages):

    pageobj = pdfreader.getPage(pagID)

    pdfWriter.addPage(pageobj)       
  
    pdfOutputFile = open(name, 'wb')
    pdfWriter.write(pdfOutputFile)
    pdfOutputFile.close()
    pdffileobj.close()

    return name



def split_pages(src, dst):
    src_f = file(src, 'r+b')
    dst_f = file(dst, 'w+b')

    input = PdfFileReader(src_f)
    output = PdfFileWriter()

    for i in range(input.getNumPages()):
        p = input.getPage(i)
        q = copy.copy(p)
        q.mediaBox = copy.copy(p.mediaBox)

        x1, x2 = p.mediaBox.lowerLeft
        x3, x4 = p.mediaBox.upperRight

        x1, x2 = math.floor(x1), math.floor(x2)
        x3, x4 = math.floor(x3), math.floor(x4)
        #x5, x6 = math.floor(x3/2.13), math.floor(x4/2.13)
        x5, x6 = math.floor(x3/1.75), math.floor(x4/1.75)

        if x3 > x4:
            # horizontal
            p.mediaBox.upperRight = (x5, x4)
            p.mediaBox.lowerLeft = (x1, x2)

            q.mediaBox.upperRight = (x3, x4)
            q.mediaBox.lowerLeft = (x5, x2)
        else:
            # vertical
            p.mediaBox.upperRight = (x3, x4)
            p.mediaBox.lowerLeft = (x1, x6)

            q.mediaBox.upperRight = (x3, x6)
            q.mediaBox.lowerLeft = (x1, x2)


        p.artBox = p.mediaBox
        p.bleedBox = p.mediaBox
        p.cropBox = p.mediaBox

        q.artBox = q.mediaBox
        q.bleedBox = q.mediaBox
        q.cropBox = q.mediaBox

        output.addPage(q)
        output.addPage(p)

    output.write(dst_f)
    src_f.close()
    dst_f.close()



def pdf_excel(pdffile):

    c = pdftables_api.Client(api_key)
    return c.xlsx(pdffile, 'demo.xlsx')


def sheet_text_exists(text,excelfile):

    wb = openpyxl.load_workbook(excelfile)    
    
    for sheets in wb.get_sheet_names():
    
        set_data_loader,counter = {},0
    
        core.info("Analysing Current Sheet - %s "%sheets)
        core.info("Searching and Filtering User-info ")
    
        currentSheet = wb.get_sheet_by_name(sheets)

        for ws in currentSheet.iter_rows():

            try:
                    #yield  [str(cell.value).strip() for cell in ws if type(cell.value) is not type(None) and 'Change' not in str(cell.value)]
                    validatedata = ''.join([str(cell.value).strip() for cell in ws if type(cell.value) is not type(None) and 'Change' not in str(cell.value)]).strip()

                    if text.strip().replace(' ','') in validatedata.strip().replace(' ',''):                        
                        
                        return sheets,currentSheet

            except UnicodeEncodeError : pass

    return False



def read_excel(text,excelFile):

    sheetName,currentsheetobject = sheet_text_exists(text=text,excelfile = excelFile)

    if bool(sheetName):

        set_data_loader,counter = {},0

        core.info("Analysing Current Sheet - %s "%sheetName)
        core.info("Searching and Filtering User-info ")

        for ws in currentsheetobject.iter_rows():
        
            try:
                
                yield  [str(cell.value).strip() for cell in ws if type(cell.value) is not type(None) and 'Change' not in str(cell.value) and 'Average' not in str(cell.value)]

            except UnicodeEncodeError : pass

    else:
        
        core.critical('Either search term is invalid or excel file path not properly entered, please  check and try again')


def analyze_rows(excelrows,start_counter=0):
    
    for values in excelrows:

        for items in filterterms : 

            if start_counter >0 and re.search(items.replace(' ','').strip(),''.join(values).replace(' ','').strip()):                
                return        
        
        if current_search_term.replace(' ','').strip().lower() in ''.join(values).replace(' ','').strip().lower():

            start_counter += 1

        if start_counter > 0:         
            
            if not values == [] :               
                yield values



def process(searchtext,results = {},excelfile = None):

    products = []
    ratings = []

    for pos,data in enumerate(analyze_rows(read_excel(searchtext,excelfile))):

            if pos == 0 : results.__setitem__('Table',''.join(data))

            elif pos == 1 :  results.__setitem__('Phase',','.join(data))
            
            else: 
 
                products.append(data.__getitem__(0))
                ratings.append(data.__getslice__(1,len(data)))

                results.__setitem__('Products',products)
                results.__setitem__('ProductRating',ratings)
      
    return results


def final(searchtext,excelfile):

        getData = process(searchtext,excelfile=excelfile)
        counter = phasecount= 0
        publish = {}
        phasesratings = []

        publish.__setitem__("Table Name : ", getData.__getitem__('Table'))

        for products in getData.__getitem__('Products'):            

            counter += 1
            phasecount= 0

            for phases in getData.__getitem__('Phase').split(','):                                                       
                    phasecount += 1
                    phasesratings.append({phases:getData.__getitem__('ProductRating')[counter-1][phasecount-1]})

            publish.__setitem__(products,phasesratings)

            phasesratings = []

        return publish

                    




'''
    Step :: 1 ==> Extract data from PDF and move to a new file if the text exists
'''

'''core.info("Step 1 ==> Extract data from PDF and move to a new file if the text exists")
_extract_data_from_pdf('testm.pdf',searchTerm)
core.info("Step 1 Completed. Please refer the directory for the o/p file! ")

'''
#print final(excelfile= 'demo.xlsx')

#print sheet_text_exists(text = searchTerm,excelfile='demo.xlsx')

'''
import pdftables_api

c = pdftables_api.Client('my-api-key')
c.xlsx('input.pdf', 'output.xlsx')'''


#print sheet_text_exists(text = searchTerm3,excelfile='testm_39.xlsx')

final(current_search_term,excelfile= 'testm_38.xlsx')